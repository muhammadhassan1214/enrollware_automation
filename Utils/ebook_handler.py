import os
import logging
from datetime import date
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from Utils.mail_sender.email_sender import send_email
from Utils.mail_sender.email_generator import generate_ebook_email

EBOOK_PRODUCT_CODES = {"25-3149"}
EBOOK_KEYWORD = "ebook"
WORKBOOK_PATH = os.path.join("data", "eBook_available_codes.xlsx")


def is_ebook_order(order: Dict[str, str]) -> bool:
    """Return True if the order represents an eBook purchase."""
    product_code = (order.get("product_code") or "").strip()
    course_name = (order.get("course_name") or "").lower()
    return product_code in EBOOK_PRODUCT_CODES or EBOOK_KEYWORD in course_name


def _parse_name(full_name: str) -> Tuple[str, str]:
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _load_sheet(logger: logging.Logger):
    if not os.path.exists(WORKBOOK_PATH):
        logger.error(f"Workbook not found at {WORKBOOK_PATH}")
        return None, None, None
    try:
        wb = load_workbook(WORKBOOK_PATH)
        ws = wb.active
    except Exception as exc:
        logger.error(f"Unable to open workbook: {exc}")
        return None, None, None

    header_row = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    required_headers = [
        "First Name", "Last Name", "Email", "Product Code", "Product", "Product Status",
        "Product URL / Code", "Group", "Assignment Date", "Completion Date", "Skills Check"
    ]
    missing = [h for h in required_headers if h not in header_row]
    if missing:
        logger.error(f"Workbook is missing required columns: {missing}")
        return None, None, None

    return wb, ws, header_row


def _get_available_rows(ws, headers: Dict[str, int]) -> List[int]:
    status_col = headers["Product Status"]
    available_rows = []
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=status_col).value
        status_text = str(status_cell).strip().lower() if status_cell is not None else ""
        if status_text.startswith("available"):
            available_rows.append(row_idx)
    return available_rows


def _collect_codes(ws, headers: Dict[str, int], rows: List[int]) -> List[str]:
    code_col = headers["Product URL / Code"]
    codes: List[str] = []
    for row_idx in rows:
        val = ws.cell(row=row_idx, column=code_col).value
        if val is None:
            continue
        codes.append(str(val).strip())
    return codes


def _update_rows(ws, headers: Dict[str, int], rows: List[int], recipient_name: str, recipient_email: str):
    first_name, last_name = _parse_name(recipient_name)
    today_str = date.today().isoformat()
    for row_idx in rows:
        ws.cell(row=row_idx, column=headers["First Name"]).value = first_name
        ws.cell(row=row_idx, column=headers["Last Name"]).value = last_name
        ws.cell(row=row_idx, column=headers["Email"]).value = recipient_email
        ws.cell(row=row_idx, column=headers["Product Status"]).value = "Assigned"
        ws.cell(row=row_idx, column=headers["Assignment Date"]).value = today_str


def process_ebook_orders(orders: List[Dict[str, str]], logger: logging.Logger) -> bool:
    """Process eBook orders: reserve codes, email them, and update the workbook."""
    if not orders:
        return False

    recipient_email = (orders[0].get("email") or "").strip()
    recipient_name = (orders[0].get("name") or "").strip() or "Customer"
    total_quantity = sum(int(order.get("quantity", 0) or 0) for order in orders)

    if not recipient_email:
        logger.error("No email found for eBook order; cannot send codes.")
        return False
    if total_quantity <= 0:
        logger.error("Invalid quantity for eBook order.")
        return False

    wb, ws, headers = _load_sheet(logger)
    if not wb:
        return False

    available_rows = _get_available_rows(ws, headers)
    if len(available_rows) < total_quantity:
        logger.error(f"Not enough eBook codes available. Needed {total_quantity}, found {len(available_rows)}")
        return False

    rows_to_use = available_rows[:total_quantity]
    codes = _collect_codes(ws, headers, rows_to_use)
    if len(codes) < total_quantity:
        logger.error("Some selected rows do not contain codes; aborting eBook processing.")
        return False

    email_body = generate_ebook_email(recipient_name, codes)
    subject = "Your Heartsaver First Aid CPR AED eBook Codes"

    if not send_email(subject, email_body, recipient_email, recipient_name):
        logger.error("Failed to send eBook email; workbook not updated.")
        return False

    _update_rows(ws, headers, rows_to_use, recipient_name, recipient_email)
    try:
        wb.save(WORKBOOK_PATH)
    except Exception as exc:
        logger.error(f"Failed to save workbook updates: {exc}")
        return False

    logger.info(f"Sent {len(codes)} eBook codes to {recipient_email} and updated the workbook.")
    return True

